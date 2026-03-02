import openpyxl
import sys

try:
    file_path = '1_StandardReport.xlsx'
    
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    
    print(f"Sheet: {ws.title}")
    
    def print_row(row_idx, row_values):
        print(f"\nRow {row_idx}:")
        for i, val in enumerate(row_values):
            if val is not None:
                print(f"  Col {i}: {val}")

    # Inspect specific rows
    rows = list(ws.iter_rows(max_row=20, values_only=True))
    
    # Header Info (Row 1-3)
    print_row(1, rows[1]) # Python 0-indexed list, Row 2 in Excel? Verify index.
    # iter_rows yields starting from row 1. list index 0 = row 1.
    
    print_row(1, rows[0])
    print_row(3, rows[2])
    print_row(10, rows[9]) # Header?
    print_row(11, rows[10]) # Data?
    print_row(12, rows[11]) # Data?

except Exception as e:
    print(f"Error: {e}")
